import argparse
import os
from glob import glob
from openpyxl import load_workbook, Workbook

CONFIG_PATH = "config/config.xlsx"
TABLES_DIR = "tables"


def load_config():
    wb = load_workbook(CONFIG_PATH)
    providers = {}
    for row in wb["providers"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            providers[row[0]] = row[1]

    product_mapping = []
    for row in wb["product_mapping"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            product_mapping.append({
                "product_name": row[0],
                "store": row[1],
                "stock_type": row[2],
                "provider": row[3],
                "provider_product_name": row[4]
            })

    desired_stock = {}
    for row in wb["desired_stock"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[2]:
            key = (row[0], row[1], row[2])
            desired_stock[key] = int(row[3]) if row[3] else 0

    return providers, product_mapping, desired_stock


def find_stock_file(store, stock_type):
    pattern = f"{stock_type}_{store}_*.xlsx"
    files = glob(os.path.join(TABLES_DIR, pattern))
    if not files:
        raise FileNotFoundError(f"No file found for {pattern}")
    files.sort(reverse=True)
    return files[0]


def parse_accessories_stock(filepath):
    products = {}
    wb = load_workbook(filepath)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        product_a = row[0]
        qty_a = row[1] if len(row) > 1 else None
        product_c = row[2] if len(row) > 2 else None

        if product_a and isinstance(product_a, str) and product_a.strip():
            if product_a.strip() != "CARGADORES" and product_a.strip() != "PARLANTES":
                qty = int(float(qty_a)) if qty_a and isinstance(qty_a, (int, float)) else 0
                products[product_a.strip()] = qty

        if product_c and isinstance(product_c, str) and product_c.strip():
            qty = int(float(qty_a)) if qty_a and isinstance(qty_a, (int, float)) else 0
            products[product_c.strip()] = qty

    return products


def parse_glass_stock(filepath):
    products = {}
    wb = load_workbook(filepath)
    if "GLASS" not in wb.sheetnames:
        return products
    ws = wb["GLASS"]

    headers = []
    for i, cell in enumerate(ws[1]):
        if cell.value:
            headers.append((i, cell.value))
        else:
            headers.append((i, None))

    for row in ws.iter_rows(min_row=2, values_only=True):
        model = row[0]
        if not model or not isinstance(model, str) or not model.strip():
            continue

        model = model.strip()
        for col_idx, header in headers[1:]:
            if header:
                brand = str(header).strip()
                qty = row[col_idx] if col_idx < len(row) else None
                qty = int(float(qty)) if qty and isinstance(qty, (int, float)) else 0
                product_name = f"{brand} {model}"
                products[product_name] = qty

    return products


def parse_stock(filepath, stock_type):
    if stock_type == "ACCESORIOS":
        return parse_accessories_stock(filepath)
    elif stock_type == "GLASS":
        return parse_glass_stock(filepath)
    else:
        raise ValueError(f"Unknown stock type: {stock_type}")


def get_missing(products, desired_stock, store, stock_type):
    missing = {}
    for key, desired_qty in desired_stock.items():
        s, st, product_name = key
        if s != store or st != stock_type:
            continue
        actual_qty = products.get(product_name, 0)
        diff = desired_qty - actual_qty
        if diff > 0:
            missing[product_name] = {
                "actual": actual_qty,
                "desired": desired_qty,
                "missing": diff
            }
    return missing


def update_wanted_file(store, stock_type, products, desired_stock):
    missing = get_missing(products, desired_stock, store, stock_type)

    filename = f"WANTED_{stock_type}_{store}.xlsx"
    filepath = os.path.join(TABLES_DIR, filename)

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        if "STOCK" in wb.sheetnames:
            ws = wb["STOCK"]
        else:
            ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "STOCK"
        ws.append(["product_name", "actual_qty", "desired_qty", "missing_qty"])

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value in missing:
            row[1].value = missing[row[0].value]["actual"]
            row[2].value = missing[row[0].value]["desired"]
            row[3].value = missing[row[0].value]["missing"]
            del missing[row[0].value]

    for product_name, data in missing.items():
        ws.append([product_name, data["actual"], data["desired"], data["missing"]])

    wb.save(filepath)
    print(f"Updated: {filepath}")
    return missing


def generate_orders(missing, product_mapping, store, stock_type):
    provider_products = {}
    for mapping in product_mapping:
        if mapping["store"] != store or mapping["stock_type"] != stock_type:
            continue
        product_name = mapping["product_name"]
        if product_name in missing:
            provider = mapping["provider"]
            provider_product_name = mapping["provider_product_name"]
            if provider not in provider_products:
                provider_products[provider] = []
            provider_products[provider].append({
                "provider_product_name": provider_product_name,
                "quantity": missing[product_name]["missing"]
            })

    for provider, items in provider_products.items():
        if not items:
            continue
        filename = f"ORDER_{stock_type}_{store}_{provider}.xlsx"
        filepath = os.path.join(TABLES_DIR, filename)
        wb = Workbook()
        ws = wb.active
        ws.append(["provider_product_name", "quantity"])
        for item in items:
            ws.append([item["provider_product_name"], item["quantity"]])
        wb.save(filepath)
        print(f"Created: {filepath}")


def main():
    parser = argparse.ArgumentParser(description="Generate stock orders")
    parser.add_argument("--store", required=True, help="Store identifier (e.g., LA_PLATA)")
    parser.add_argument("--stock_type", required=True, choices=["ACCESORIOS", "GLASS"],
                        help="Stock type")
    args = parser.parse_args()

    providers, product_mapping, desired_stock = load_config()

    stock_file = find_stock_file(args.store, args.stock_type)
    print(f"Reading stock from: {stock_file}")

    products = parse_stock(stock_file, args.stock_type)
    print(f"Found {len(products)} products in stock file")

    missing = update_wanted_file(args.store, args.stock_type, products, desired_stock)
    print(f"Found {len(missing)} products with missing stock")

    if missing:
        generate_orders(missing, product_mapping, args.store, args.stock_type)


if __name__ == "__main__":
    main()
